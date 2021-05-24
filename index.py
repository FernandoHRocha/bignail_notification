from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium import webdriver
from twilio.rest import Client
import openpyxl
import dados
import time
from openpyxl import Workbook

#CONFIGURAÇÃO DE VARIAVEIS
#PLANILHA DE CONTROLE
pregao_path = dados.pregao_path
excel_read = openpyxl.load_workbook(pregao_path)
#TWILIO
twilio_account = dados.twilio_account
twilio_pass = dados.twilio_pass
twilio_from_number=dados.twilio_from_number
twilio_to_number=dados.twilio_to_number
client = Client(twilio_account, twilio_pass)
#PREGAO
pregao_account=dados.pregao_account
pregao_pass=dados.pregao_pass
pregao_address=dados.pregao_address
#pregao_sheet=xlrd.open_workbook('notificacoes.xlsx')
#SELENIUM
chrome_options = Options()
chrome_options.headless = True
sel_driver = webdriver.Chrome("chromedriver.exe",chrome_options=chrome_options)
sel_driver.maximize_window()
sel_driver.get(pregao_address)
sel_delay=0.2
#message = client.messages.create(body='Hello there!', from_=twilio_from_number, to=twilio_to_number)
#print(message.sid)

def twilio_msg(msg):
    client.messages.create(body=msg, from_=twilio_from_number, to=twilio_to_number)
def sel_enterField(path, text):
    field = WebDriverWait(sel_driver,10).until(expected_conditions.presence_of_element_located((By.XPATH,path)))
    field.send_keys(text)
    time.sleep(sel_delay)
def sel_enterFieldElement(element, text):
    element.send_keys(text)
    time.sleep(sel_delay)
def sel_buttonClick(path):
    button = WebDriverWait(sel_driver,10).until(expected_conditions.element_to_be_clickable((By.XPATH,path)))
    button.click()
def sel_switchFrame(path):
    sel_driver.switch_to.default_content()
    frame = WebDriverWait(sel_driver,10).until(expected_conditions.frame_to_be_available_and_switch_to_it((By.XPATH,path)))
def sel_mouseHover(path):
    clickable = WebDriverWait(sel_driver,10).until(expected_conditions.element_to_be_clickable((By.XPATH,path)))
    hover = ActionChains(sel_driver).move_to_element(clickable)
    hover.perform()
def sel_getElement(path):
    el = WebDriverWait(sel_driver,10).until(expected_conditions.presence_of_element_located((By.XPATH,path)))
    return el
def sel_getElements(path):
    el = WebDriverWait(sel_driver,10).until(expected_conditions.presence_of_all_elements_located((By.XPATH,path)))
    return el
def sel_newWindowClick(path):
    action = ActionChains(sel_driver).key_down(Keys.SHIFT)
    action.perform()
    action = ActionChains(sel_driver).click(path)
    action.perform()
    action = ActionChains(sel_driver).key_up(Keys.SHIFT)
    action.perform()

#ACESSAR O SISTEMA
sel_buttonClick('//*[@id="card0"]/div/div/div/div[2]/button')
sel_enterField('//*[@id="txtLogin"]',pregao_account)
sel_enterField('//*[@id="txtSenha"]', pregao_pass)
sel_buttonClick('//*[@id="card0"]/div/div/div[2]/div[4]/button[2]')
print('LOGADO COM SUCESSO')
#FECHAR POPUP
sel_mainWindow = sel_driver.window_handles[0]
time.sleep(0.5)
sel_windowToClose = sel_driver.window_handles[1]
sel_driver.switch_to.window(sel_windowToClose)
sel_driver.close()
sel_driver.switch_to.window(sel_mainWindow)
#ACESSAR SEÇÃO DE PREGÕES EM ANDAMENTO
sel_switchFrame('/html/frameset/frame[1]')
sel_mouseHover('/html/body/div[2]/div[1]')
sel_switchFrame('//*[@id="corpo"]/frame')
time.sleep(0.2)
sel_buttonClick('/html/body/div[1]/div[4]')
sel_buttonClick('/html/body/div[1]/ul/li[4]/a')

link = []
nail = []
def sel_refreshTable():
    sel_switchFrame('/html/frameset/frameset/frame')
    table = sel_getElements('/html/body/div[1]/table/tbody/tr[5]/td[2]/table/tbody/*')
    link = []
    nail = []
    for index in range(0,len(table)):
        if( index > 0 ):
            linha = table[index].find_elements_by_xpath('./*')
            link.append(linha[0])
            aux_nail=[]
            aux_nail.append(linha[1].text)
            aux_nail.append(linha[2].text)
            aux_nail.append(linha[3].text)
            nail.append(aux_nail)

sel_switchFrame('/html/frameset/frameset/frame')
table = sel_getElements('/html/body/div[1]/table/tbody/tr[5]/td[2]/table/tbody/*')
link = []
nail = []
for index in range(0,len(table)):
    if( index > 0 ):
        linha = table[index].find_elements_by_xpath('./*')
        link.append(linha[0])
        aux_nail=[]
        aux_nail.append(linha[1].text)
        aux_nail.append(linha[2].text)
        aux_nail.append(linha[3].text)
        nail.append(aux_nail)

def sel_lerPregao(linka, sheet):
    sel_mainWindow = sel_driver.window_handles[0]
    sel_newWindowClick(linka)
    sel_windowToClose = sel_driver.window_handles[1]
    sel_driver.switch_to.window(sel_windowToClose)
    sel_buttonClick('/html/body/div[1]/table/tbody/tr[6]/td[2]/input')
    sel_driver.close()
    sel_windowToClose = sel_driver.window_handles[1]
    sel_driver.switch_to.window(sel_windowToClose)
    table = sel_getElements('/html/body/table[2]/tbody/*')
    sh = excel_read[sheet]
    for index in range(0,len(table)):
        #if(index<50):
        msg = table[index].find_elements_by_xpath('./*')
          #  cl = sh.cell(row=index+1, column=1)
           # cl.value = msg[0].text
            #cl = sh.cell(row=index+1, column=2)
            #cl.value = msg[1].text
        if(index==0):
            twilio_msg(sheet+" "+ msg[0].text + " " + msg[1].text)
    excel_read.save(pregao_path)
    sel_driver.close()
    sel_driver.switch_to.window(sel_mainWindow)

#FUNÇÕES PLANILHA
for index in range(0,len(nail)):
    sel_refreshTable()
    criar = True
    sheet_name=str(nail[index][0]+" | "+nail[index][1])
    for sheet in range(0,len(excel_read.sheetnames)):
        if(excel_read.worksheets[sheet].title == sheet_name):
            criar = False
            if(link[index].text == 'Acompanhar'):
                sel_lerPregao(link[index], sheet_name)
    if(criar):
        excel_read.create_sheet(sheet_name)
        excel_read.save(pregao_path)
        if(link[index].text == 'Acompanhar'):
            for sheet in range(0,len(excel_read.sheetnames)):
                if(excel_read.worksheets[sheet].title == sheet_name):
                    sel_lerPregao(link[index], sheet_name)
sel_driver.quit()
excel_read.save(pregao_path)
excel_read.close()