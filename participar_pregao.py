from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium import webdriver
from openpyxl import Workbook
import dados_cotacao
import openpyxl
import dados
import time
import sys

#CONFIGURAÇÃO DE VARIAVEIS
#PLANILHA DE CONTROLE
pregao_path = dados.pregao_path
excel_read = openpyxl.load_workbook(pregao_path)
#PREGAO
pregao_account=dados.pregao_account
pregao_pass=dados.pregao_pass
pregao_address=dados.pregao_address
#pregao_sheet=xlrd.open_workbook('notificacoes.xlsx')
#SELENIUM
options = webdriver.ChromeOptions()
#options.add_argument ('--headless')
sel_driver = webdriver.Chrome("chromedriver.exe", options=options)
sel_driver.maximize_window()
sel_driver.get(pregao_address)
sel_delay=0.2

def sel_enterField(path, text):
    field = WebDriverWait(sel_driver,10).until(expected_conditions.presence_of_element_located((By.XPATH,path)))
    field.send_keys(text)
    time.sleep(sel_delay)
def sel_enterFieldElement(element, text):
    element.send_keys(text)
    time.sleep(sel_delay)
def sel_buttonClick(path):
    button = WebDriverWait(sel_driver,10).until(expected_conditions.element_to_be_clickable((By.XPATH,path)))
    button.click()
def sel_switchFrame(path):
    sel_driver.switch_to.default_content()
    frame = WebDriverWait(sel_driver,10).until(expected_conditions.frame_to_be_available_and_switch_to_it((By.XPATH,path)))
def sel_mouseHover(path):
    clickable = WebDriverWait(sel_driver,10).until(expected_conditions.element_to_be_clickable((By.XPATH,path)))
    hover = ActionChains(sel_driver).move_to_element(clickable)
    hover.perform()
def sel_getElement(path):
    el = WebDriverWait(sel_driver,10).until(expected_conditions.presence_of_element_located((By.XPATH,path)))
    return el
def sel_getElements(path):
    el = WebDriverWait(sel_driver,10).until(expected_conditions.presence_of_all_elements_located((By.XPATH,path)))
    return el
def sel_newWindowClick(path):
    action = ActionChains(sel_driver).key_down(Keys.SHIFT)
    action.perform()
    action = ActionChains(sel_driver).click(path)
    action.perform()
    action = ActionChains(sel_driver).key_up(Keys.SHIFT)
    action.perform()

def disputar(sel_link):
    sel_link.click()
    sel_mainWindow = sel_driver.window_handles[0]
    sel_windowDisputa = sel_driver.window_handles[1]
    sel_driver.switch_to.window(sel_windowDisputa)
    

##LEITURA DOS DADOS DA PLANILHA
# workbook = openpyxl.load_workbook('pregao.xlsx').sheet_by_name('Planilha1')
# itens=[]
# for row in range(3,workbook.nrows):
#     rowItens=[]
#     for col in range(0,6):
#         if(col == 0 or col == 4):
#             item = int(workbook.cell(int(row),int(col)).value)
#             rowItens.append(item)
#         else:
#             rowItens.append(workbook.cell(int(row),int(col)).value)
#     itens.append(rowItens)
#ACESSAR O SISTEMA
sel_buttonClick('//*[@id="card0"]/div/div/div/div[2]/button')
sel_enterField('//*[@id="txtLogin"]',pregao_account)
sel_enterField('//*[@id="txtSenha"]', pregao_pass)
sel_buttonClick('//*[@id="card0"]/div/div/div[2]/div[4]/button[2]')
print('Logado no sistema ComprasNet')
#FECHAR POPUP
sel_mainWindow = sel_driver.window_handles[0]
time.sleep(0.5)
sel_windowToClose = sel_driver.window_handles[1]
sel_driver.switch_to.window(sel_windowToClose)
sel_driver.close()
sel_driver.switch_to.window(sel_mainWindow)
#ACESSAR SEÇÃO DE PREGÕES EM DISPUTA
sel_switchFrame('/html/frameset/frame[1]')
sel_mouseHover('/html/body/div[2]/div[1]')
sel_switchFrame('//*[@id="corpo"]/frame')
time.sleep(0.2)
try:
    sel_buttonClick('/html/body/div[1]/div[4]')
except:
    sel_switchFrame('/html/frameset/frame[1]')
    sel_mouseHover('/html/body/div[2]/div[1]')
time.sleep(0.2)
sel_switchFrame('//*[@id="corpo"]/frame')
sel_buttonClick('/html/body/div[1]/ul/li[2]/a')
#sel_buttonClick('/html/body/div[1]/ul/li[2]/a')
table = sel_getElement('/html/body/table/tbody/tr[2]/td/table[2]/tbody/tr[3]/td[2]/table/tbody')
rows = table.find_elements_by_xpath('./*')
infos = rows[1].find_elements_by_xpath('./*')
if(len(infos) == 1):
    print('Nenhum pregão está em fase de disputa de lances no momento.')
else:
    print('Temos 1 pregão em andamento do orgão: ' + infos[3].text + ', nº ' + infos[1].text + ' UASG: ' + infos[2].text)
    disputar(infos[0])