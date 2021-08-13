import openpyxl
#LEITURA DOS DADOS DA PLANILHA
wb = openpyxl.load_workbook('COTACAO.xlsx', data_only=True)['Planilha1']


def import_numero():
    numero = wb.cell(1,2).value
    return numero

def import_uasg():
    uasg = wb.cell(1,4).value
    return uasg

def import_items():
    itens=[]
    for row in range(3,wb.max_row):
        rowItens=[]
        for col in range(1,wb.max_column):
            if (col == 12):
                rowItens.append(wb.cell(row,col).value)
            if (col <= 6):
                if(col != 2 and col !=3):
                    rowItens.append(wb.cell(row,col).value)
                if(col == 5):
                    total = wb.cell(row,4).value * wb.cell(row,5).value
                    rowItens.append(total)
        itens.append(rowItens)
    return itens