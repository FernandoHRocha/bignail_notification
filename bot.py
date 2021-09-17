import sel_operacoes_comum as sel
import openpyxl
import time
import os
import sys
import traceback

sel_delay=0.5
sel_driver = ''
caminho_pasta = 'C:/Fernando/LOJA/outros/twilio/bignail_notification/COTACAO'
arquivo_planilha = 'COTACAO.xlsx'
arquivo_proposta = 'PROPOSTA.pdf'
arquivo_documentacao = 'DOCUMENTACAO.rar'
endereco_planilha = caminho_pasta + '/' + arquivo_planilha
endereco_proposta = caminho_pasta + '/' + arquivo_proposta
endereco_documentacao = caminho_pasta + '/' + arquivo_documentacao
itens_aguardando_disputa = 'Aguardando disputa'
itens_fase_disputa = 'Em disputa'
itens_disputa_encerrados = 'Encerrados'
item_etapa_aberta = 'Etapa aberta'
item_etapa_fechada = 'Etapa fechada'
tempo_prorrogado = 'prorrogacao'

def abrir_pasta():
    path = os.path.realpath(caminho_pasta)
    os.startfile(path)

def oferecer_abrir_pasta():#OFERECE A OPÇÃO DE ABRIR A PASTA QUE DEVERÁ CONTER AS PLANILHAS CONTROLE E COTAÇÃO
    print('O nome dos arquivos devem ser:\nPlanilha de cotação "'+arquivo_planilha+'"')
    print('Proposta assinada "'+arquivo_proposta+'"')
    print('Documentação compactada "'+arquivo_documentacao+'"')
    print('1 - Abrir a pasta.')
    print('2 - A planilha já está na pasta.')
    escolha = input('>')
    if(escolha == '1'):
        abrir_pasta()
        print('Aguardando para continuar.')
        escolha = input('>')

def converter_texto_para_decimal(texto):
    texto = float(texto.replace("R$","").strip().replace(".","").replace(",","."))
    return "{:.2f}".format(texto)

def converter_intervalo_minimo(intervalo):
    padrao1 = 'Intervalo mínimo entre lances: R$ '
    padrao2 = 'Não há intervalo mínimo entre lances'
    padrao3 = 'Intervalo mínimo entre lances:'
    if padrao1 in intervalo:
        intervalo = intervalo.replace(padrao1,"")
        return converter_texto_para_decimal(intervalo)
    elif padrao2 in intervalo:
        return 0.10
    elif '%' in intervalo:
        intervalo = intervalo.replace(padrao3,"").replace('%','')
        return converter_texto_para_decimal(intervalo)#APLICAR LÓGICA PARA INTERVALOS PERCENTUAIS

def converter_tempo_restante(tempo):
    tempo = tempo.split(":")
    return int(tempo[1])+(int(tempo[0])*60)

class ComprasNet:#LEVA A APLICAÇÃO ATÉ UM LUGAR EM COMUM DENTRO DO COMPRASNET E MOSTRA AS OPÇÕES DE OPERAÇÕES

    def iniciar(self):
        sel.configurar_webdriver(self,True)
        sel.coletar_credenciais_acessar_sistema(self)
        sel.acessar_menu_comprasnet(self)
        self.oferecer_opcoes()

    def oferecer_opcoes(self):
        class_dict={'1':Registrar,'2':Disputar}
        print('Escolha o modo de operação')
        print('1 - Registrar proposta.')
        print('2 - Participar da disputa de lances')
        escolha = input('>')
        global sel_driver
        sel_driver = self.sel_driver
        class_dict[escolha].iniciar(class_dict[escolha])

class Registrar:#REGISTRA O PREGÃO REFERENTE AO ARQUIVO DE COTAÇÃO

    def iniciar(self):
        self.sel_driver = sel_driver
        oferecer_abrir_pasta()
        self.ler_planilha_cotacao(self)
        self.acessar_cadastro(self)
        self.submeter_docs = True
        if(len(self.itens_cotacao)>0):
            print('iniciando processo de registro de item')
            while(True):
                print('Temos ',len(self.itens_cotacao),' item(s) para registrar.')
                self.encontrar_proximo_item_cotado(self,self.itens_cotacao[0],self.identificar_pagina_registro(self))
                if(len(self.itens_cotacao)==0):
                    break
        else:
            print('A planilha não possui itens cotados, ou não foi possível a sua identificação.')
        self.aceitar_insercao_item(self,False)
    
    def ler_planilha_cotacao(self):
        wb = openpyxl.load_workbook('./COTACAO/COTACAO.xlsx', data_only=True)['Controle']
        self.pregao = str(wb.cell(2,1).value)
        self.uasg = str(wb.cell(2,2).value)
        self.abertura = str(wb.cell(2,3).value)
        self.hora = str(wb.cell(2,4).value)
        self.inserir_orgao = str(wb.cell(2,5).value)
        wb = openpyxl.load_workbook('./COTACAO/COTACAO.xlsx', data_only=True)['Planilha1']
        itens=[]
        for row in range(2,wb.max_row):
            auxiliar_item_linha=[]
            colunas_interesse=[1,2,3,4,5]
            colunas_monetarias =[4]
            for col in colunas_interesse:
                if(col in colunas_monetarias):
                    valor = str(round(wb.cell(row,col).value,2))
                    if(len(valor.split('.'))<2):
                        valor = valor + '.00'
                    elif(len(valor.split('.')[1])<2):
                        valor = valor.split('.')[0]+'.'+valor.split('.')[1]+'0'
                    auxiliar_item_linha.append(str(valor).replace('.',','))
                else:
                    auxiliar_item_linha.append(str(wb.cell(row,col).value))
            valor = str(round(round(wb.cell(row,4).value,2)*int(wb.cell(row,5).value),2))
            if(len(valor.split('.'))<2):
                    valor = valor + '.00'
            elif(len(valor.split('.')[1])<2):
                valor = valor.split('.')[0]+'.'+valor.split('.')[1]+'0'
            auxiliar_item_linha.append(str(valor).replace('.',','))
            itens.append(auxiliar_item_linha)
        self.itens_cotacao = itens
        #print(self.itens_cotacao)
        # self.itens_cotacao[0] IDENTIFICADOR
        # self.itens_cotacao[1] DESCRIÇÃO
        # self.itens_cotacao[2] MARCA
        # self.itens_cotacao[3] VALOR UNITÁRIO
        # self.itens_cotacao[4] QUANTIDADE OFERTADA
        # self.itens_cotacao[5] QUANTIDADE TOTAL

        while(True):
            print('O cadatro será realizado para o pregão: '+self.pregao+' do uasg: '+self.uasg)
            print('1 - Continuar.')
            print('2 - Abrir pasta de cotação para alterar a planilha de cotação.')
            choose = input('>')
            if(choose == '1'):
                break
            elif(choose == '2'):
                abrir_pasta()
                input('>')
                self.ler_planilha_cotacao()
                break

    def acessar_cadastro(self):
        sel.clicar_xpath(self,'/html/body/div[1]/ul/li[1]/a')
        sel.clicar_xpath(self,'/html/body/div[1]/ul/li[1]/span')
        sel.enterField(self,'/html/body/form/table/tbody/tr[2]/td/table[2]/tbody/tr[4]/td[2]/table/tbody/tr/td/table/tbody/tr[3]/td[2]/input',self.uasg)
        sel.enterField(self,'/html/body/form/table/tbody/tr[2]/td/table[2]/tbody/tr[4]/td[2]/table/tbody/tr/td/table/tbody/tr[4]/td[2]/input',self.pregao)
        sel.clicar_xpath(self,'/html/body/form/table/tbody/tr[2]/td/table[2]/tbody/tr[4]/td[2]/table/tbody/tr/td/table/tbody/tr[7]/td/input[3]')
        sel.clicar_xpath(self,'/html/body/table/tbody/tr[2]/td/table[2]/tbody/tr[2]/td[2]/form/table/tbody/tr[2]/td/table/tbody/tr[2]/td[1]/a')

    def identificar_pagina_registro(self):
        sel.trocar_frame(self,'/html/frameset/frameset/frame')
        tabela = sel.obter_elementos_xpath(self,'/html/body/center/table[2]/tbody/tr[4]/td/center[2]/table/tbody/tr')
        del tabela[0]
        item_registrar = []
        novo_grupo = True
        itens_pagina = []
        aux_item =[]
        for tr in tabela:
            if(novo_grupo):
                if(tr.find_element_by_xpath('./td[2]').text != ''):
                    aux_item.append(tr)
                    novo_grupo = False
                else:
                    aux_item.append(tr)
                    novo_grupo = False
            else:
                if(tr.find_element_by_xpath('./td[2]').text == ''):
                    aux_item.append(tr)
                else:
                    novo_grupo = True
                    itens_pagina.append(aux_item)
                    aux_item = []
                    aux_item.append(tr)
        itens_pagina.append(aux_item)
        return itens_pagina

    def encontrar_proximo_item_cotado(self, item_cotado, itens_pagina):
        print('Procurando inserir o item ',item_cotado[0])
        for item_pagina in itens_pagina:
            sel.trocar_frame(self,'/html/frameset/frameset/frame')
            if(item_pagina[0].find_element_by_xpath('./td[2]').text == item_cotado[0]):
                self.preencher_item_registrar(self,item_cotado = self.itens_cotacao.pop(0), item_pagina = item_pagina)
                if(self.submeter_docs):
                    self.submeter_documentacao(self)
                    self.submeter_docs = False
                break
            if(item_pagina == itens_pagina[(len(itens_pagina)-1)]):
                self.aceitar_insercao_item(self,True)
    
    def submeter_documentacao(self):
        sel.clicar_xpath(self,'/html/body/center/table[2]/tbody/tr[3]/td/table/tbody/tr[4]/td/table/tbody/tr/td/center/input[1]')
        tabela = sel.obter_elementos_xpath(self,'//table[@id="declaraEdital"]//following-sibling::table')
        for linha in tabela:
            linha.find_element_by_xpath('./tbody/tr/td[2]/input').click()
        #DOCUMENTACAO
        janela_atual = self.sel_driver.window_handles[0]
        sel.obter_elemento_id(self,'incluiAnexo').click()
        time.sleep(sel_delay)
        janela_documento = self.sel_driver.window_handles[1]
        self.sel_driver.switch_to.window(janela_documento)
        sel.enterField(self,'/html/body/form/table/tbody/tr[2]/td/table/tbody/tr[2]/td/table/tbody/tr[1]/td/input',endereco_documentacao)
        sel.clicar_xpath(self,'/html/body/form/table/tbody/tr[2]/td/table/tbody/tr[2]/td/table/tbody/tr[3]/td/input')
        sel.aceitar_alerta(self.sel_driver)
        self.sel_driver.switch_to.window(janela_atual)
        sel.trocar_frame(self,'/html/frameset/frameset/frame')
        while(True):
            time.sleep(sel_delay)
            if(len(self.sel_driver.window_handles) == 1):
                break
        #PROPOSTA
        sel.obter_elemento_id(self,'incluiAnexoP').click()
        time.sleep(sel_delay)
        janela_documento = self.sel_driver.window_handles[1]
        self.sel_driver.switch_to.window(janela_documento)
        sel.enterField(self,'/html/body/form/table/tbody/tr[2]/td/table/tbody/tr[2]/td/table/tbody/tr[1]/td/input',endereco_proposta)
        sel.clicar_xpath(self,'/html/body/form/table/tbody/tr[2]/td/table/tbody/tr[2]/td/table/tbody/tr[3]/td/input')
        sel.aceitar_alerta(self.sel_driver)
        while(True):
            time.sleep(sel_delay)
            if(len(self.sel_driver.window_handles) == 1):
                break
        self.sel_driver.switch_to.window(janela_atual)
        sel.trocar_frame(self,'/html/frameset/frameset/frame') 
        time.sleep(1)
        sel.enter_alerta(self.sel_driver)

    def preencher_item_registrar(self,item_pagina,item_cotado):
        for linha in item_pagina:
            for entrada in linha.find_elements_by_tag_name('input'):#ENTRADAS DE DADOS REFERENTES A QUANTIDADE, VALORES, MODELO
                if(entrada.is_displayed() and entrada.is_enabled()):
                    if(str(entrada.get_attribute('name')) == 'qtdOfertada'):
                        sel.enterFieldElement(entrada,item_cotado[4])
                    if(str(entrada.get_attribute('name')) == 'valorunit'):
                        sel.enterFieldElement(entrada,item_cotado[3])
                    if(str(entrada.get_attribute('name')) == 'valorprp'):
                        sel.enterFieldElement(entrada,item_cotado[5])
                    if(str(entrada.get_attribute('name')) == 'MarcaFornec'):
                        sel.enterFieldElement(entrada,item_cotado[2])
                    if(str(entrada.get_attribute('name')) == 'FabriFornec'):
                        sel.enterFieldElement(entrada,item_cotado[2])
                    if(str(entrada.get_attribute('name')) == 'ModVerFornec'):
                        sel.enterFieldElement(entrada,item_cotado[2])
            for entrada in linha.find_elements_by_tag_name('textarea'):#ENTRADAS DE DADOS REFERENTES A DESCRIÇÃO
                if(entrada.is_displayed() and entrada.is_enabled()):
                    if(str(entrada.get_attribute('name')) == 'DescrFornec'):
                        sel.enterFieldElement(entrada,item_cotado[1])
        sel.obter_elemento_id(self,'declaraEdital').find_element_by_xpath('./tbody/tr/td[2]/input').click()

    def aceitar_insercao_item(self, continuar):
        sel.obter_elemento_id(self,'incluir').click()
        time.sleep(1)
        sel.aceitar_alerta(self.sel_driver)
        if(continuar):
            try:
                sel.aceitar_alerta(self.sel_driver)
            except:
                time.sleep(2)
            sel.trocar_frame(self,'/html/frameset/frameset/frame')
            sel.obter_elemento_id(self,'proximas').click()
        else:
            print('Seus itens foram inseridos com sucesso.')

class Disputar:#DISPUTA OS PREÇOS DO PREGÃO REFERENTE AO ARQUIVO DE COTAÇÃO

    def iniciar(self):
        self.sel_driver = sel_driver
        oferecer_abrir_pasta()
        self.ler_planilha_cotacao(self)
        if(self.abrir_disputa(self)):
            self.reconhecer_disputa(self)
        return

    def ler_planilha_cotacao(self):#OBTEM UMA LISTA DE VALORES NUMÉRICOS PARA CADA ITEM COTADO
        wb = openpyxl.load_workbook('./COTACAO/COTACAO.xlsx', data_only=True)['Controle']
        self.pregao = str(wb.cell(2,1).value)
        self.uasg = str(wb.cell(2,2).value)
        wb = openpyxl.load_workbook('./COTACAO/COTACAO.xlsx', data_only=True)['Planilha1']
        self.itens=[]
        for row in range(2,wb.max_row):
            rowItens=[]
            colunas_interesse=[1,4,5,14]
            colunas_monetarias =[4,14]
            for col in colunas_interesse:
                if(col in colunas_monetarias):
                    rowItens.append(str(round(wb.cell(row,col).value,2)))
                else:
                    rowItens.append(str(wb.cell(row,col).value))
            self.itens.append(rowItens)

    def abrir_disputa(self):
        sel.clicar_xpath(self,'/html/body/div[1]/ul/li[2]/a')
        tabela = sel.obter_elementos_xpath(self,'/html/body/table/tbody/tr[2]/td/table[2]/tbody/tr[3]/td[2]/table/tbody/tr')
        del tabela[0]
        if(len(tabela[0].find_elements_by_xpath('./*'))>1):
            for linha in tabela:
                colunas = linha.find_elements_by_xpath('./td')
                codpregao = colunas[1].text
                coduasg = colunas[2].text
                if((codpregao == self.pregao) and (coduasg == self.uasg)):#SOMENTE ABRIRÁ A DISPUTA CASO CORRESPONDA AO PREGÃO E AO UASG
                    colunas[0].click()
            time.sleep(5)
            self.janela_pregoes = self.sel_driver.window_handles[0]
            self.janela_disputa = self.sel_driver.window_handles[1]
            self.sel_driver.switch_to.window(self.janela_disputa)
            return True
        else:
            print('Não foram encontradas disputas em andamento.')
            return False

    ##---------------------------------------------reconhecimento individual dos itens
    def obter_nosso_valor_aguardando(self, item):
        return converter_texto_para_decimal(str(item.find_element_by_xpath('./div[2]/div[1]/div[2]/div/div[1]/div[2]/div[2]').text))
    
    def obter_codigo_item_aguardando(self,item):
        return str(item.find_element_by_xpath('./div[1]/div[2]/div[1]/div[1]/span[1]').text)
    
    def obter_itens_tabela_aguardando(self):
        return sel.obter_elementos_xpath(self,'/html/body/app-root/div/div/div/app-cabecalho-disputa-fornecedor/div[5]/div[2]/app-disputa-fornecedor/div/p-tabview/div/div/p-tabpanel[1]/div/app-disputa-fornecedor-itens/div/p-dataview/div/div[2]/div/div')
    
    def obter_itens_tabela_disputa(self):
        return sel.obter_elementos_xpath(self,'/html/body/app-root/div/div/div/app-cabecalho-disputa-fornecedor/div[5]/div[2]/app-disputa-fornecedor/div/p-tabview/div/div/p-tabpanel[2]/div/app-disputa-fornecedor-itens/div/p-dataview/div/div[2]/div/div')
    
    def obter_codigo_item_disputa(self,item):
        return str(item.find_element_by_xpath('./div[1]/div[1]/div[1]/div/span[1]').text)
    
    def obter_etapa_item_disputa(self,item):
        etapa = str(item.find_element_by_xpath('./div[1]/div[1]/div[2]').text)
        if(item_etapa_aberta in etapa):
            return item_etapa_aberta
        else:
            return item_etapa_fechada
    
    def obter_atual_valor_disputa(self,item):
        return converter_texto_para_decimal(str(item.find_element_by_xpath('./div[2]/div[1]/div[2]/div/div[1]/div[2]/div[1]').text))
    
    def obter_nosso_valor_disputa(self,item):
        return converter_texto_para_decimal(str(item.find_element_by_xpath('./div[2]/div[1]/div[2]/div/div[1]/div[2]/div[2]').text))
    
    def obter_tempo_restante_disputa(self,item):
        try:
            tempo = converter_tempo_restante(str(item.find_element_by_xpath('./div[1]/div[2]/div/div[2]/span/span').text))
        except:
            tempo = tempo_prorrogado
        return tempo
    
    def obter_intervalo_lances_disputa(self,item):
        return converter_intervalo_minimo(str(item.find_element_by_xpath('./div[2]/div[1]/div[2]/div/div[2]/div[2]/div/div[2]/span/small').text))
    
    def obter_entrada_lance_disputa(self,item):
        return item.find_element_by_xpath('./div[2]/div[1]/div[2]/div/div[2]/div[2]/div/div[1]/input')
    
    def obter_botao_lance_disputa(self,item):
        return item.find_element_by_xpath('./div[2]/div[1]/div[2]/div/div[2]/div[2]/div/div[1]/div/button/u')
    
    def obter_menu_lances_disputa(self,item):
        try:
            botao = item.find_element_by_xpath('.//button[@title="Mostrar propostas/lances do item"]')
        except:
            botao = item.find_element_by_xpath('.//button[@title="Ocultar propostas/lances do item"]')
        return botao
    
    def obter_informacoes_item_disputa(self, item, etapa):
        if(etapa == item_etapa_aberta):
            return {
                'webelement' : item,
                'codigo_item' : self.obter_codigo_item_disputa(self,item),
                'etapa_disputa' : etapa,
                'atual_valor' : self.obter_atual_valor_disputa(self,item),
                'nosso_valor' : self.obter_nosso_valor_disputa(self,item),
                'tempo_restante' : self.obter_tempo_restante_disputa(self,item),
                'intervalo_lances' : self.obter_intervalo_lances_disputa(self,item),
                'entrada_lance' : self.obter_entrada_lance_disputa(self,item),
                'botao_lance' : self.obter_botao_lance_disputa(self,item),
                'menu_lances' : self.obter_menu_lances_disputa(self,item),
            }
        else:
            return {
                'webelement' : item,
                'codigo_item' : self.obter_codigo_item_disputa(self,item),
                'etapa_disputa' : etapa,
                'atual_valor' : self.obter_atual_valor_disputa(self,item),
                'nosso_valor' : self.obter_nosso_valor_disputa(self,item),
                'tempo_restante' : self.obter_tempo_restante_disputa(self,item),
                'intervalo_lances' : self.obter_intervalo_lances_disputa(self,item),
                'entrada_lance' : self.obter_entrada_lance_disputa(self,item),
                'botao_lance' : self.obter_botao_lance_disputa(self,item),
            }
    ##---------------------------------------------reconhecimento individual dos itens
    ##---------------------------------------------reconhecimento de itens da página de disputa
    def obter_botao_retirar_encerrados_disputa(self):
        sel.clicar_xpath(self,'/html/body/app-root/div/div/div/app-cabecalho-disputa-fornecedor/div[5]/div[2]/app-disputa-fornecedor/div/p-tabview/div/div/p-tabpanel[2]/div/app-disputa-fornecedor-itens/div/div/div[2]/div/div[1]/button').click()
        return
    ##---------------------------------------------reconhecimento de itens da página de disputa
    
    def reconhecer_valor_disputa(self,modo):#RECONHECER O VALOR DE DISPUTA GLOBAL OU UNITARIO
        if(modo == 'aguardando'):
            self.navegacao_itens[0].click()
            tabela = self.obter_itens_tabela_aguardando(self)
            item = self.obter_codigo_item_aguardando(self,tabela[0])
            valor = self.obter_nosso_valor_aguardando(self,tabela[0])
        else:
            self.navegacao_itens[1].click()
            tabela = self.obter_itens_tabela_disputa(self)
            for itens in tabela:
                item = self.obter_codigo_item_disputa(self,itens)
                valor = self.obter_nosso_valor_disputa(self,itens)
                break
        for cotado in self.itens:
            if(cotado[0] == item):
                if(str(cotado[1])==valor):
                    self.valor_disputa = 'unitario'
                    break
                else:
                    self.valor_disputa = 'global'#converter os valores mínimos para globais
                    break

    def reconhecer_disputa(self):#COLOCAR UMA ESTRUTURA DE REPETIÇÃO PARA CONTINUAR ATÉ QUE A DISPUTA ENCERRE
        time.sleep(5)
        self.modo_disputa = sel.obter_elemento_xpath(self,'/html/body/app-root/div/div/div/app-cabecalho-disputa-fornecedor/div[4]/div[1]/app-identificacao-compra/div/span').text
        self.navegacao_itens = sel.obter_elementos_xpath(self,'/html/body/app-root/div/div/div/app-cabecalho-disputa-fornecedor/div[5]/div[2]/app-disputa-fornecedor/div/p-tabview/div/ul/li')
        finalizado = False
        try:
            sel.clicar_xpath(self,'/html/body/modal-container/div/div/app-dialog-confirmacao/div/div/div[3]/div/div/button')
            finalizado = True
        except:
            pass
        
        if(finalizado == False):
            while(True):#CICLO DE REPETIÇÃO ATÉ QUE NÃO HAJA ITENS EM DISPUTA E AGUARDANDO
                print('while')
                if(self.navegacao_itens[0].text != itens_aguardando_disputa or self.navegacao_itens[1].text != itens_fase_disputa):
                    if(self.navegacao_itens[0].text != itens_aguardando_disputa):
                        self.reconhecer_valor_disputa(self,'aguardando')
                    else:
                        self.reconhecer_valor_disputa(self,'disputa')
                    if(self.navegacao_itens[1].text != itens_fase_disputa):
                        self.navegacao_itens[1].click()
                        self.reconhecer_itens_disputa(self)#COMEÇAR O CICLO DE DISPUTA DE LANCES
                    elif(self.navegacao_itens[0].text != itens_aguardando_disputa):
                        self.navegacao_itens[0].click()
                        aguardando = True
                        while(aguardando):#CICLO DE ESPERA ATÉ QUE ALGUM ITEM AGUARDANDO ENTRE EM FASE DE LANCES
                            print('Estamos aguardando algum item entrar em fase de disputa.')
                            if(self.navegacao_itens[1].text != itens_fase_disputa):
                                aguardando=False
                            else:
                                time.sleep(5)
                else:
                    self.extrair_relatorio(self)
                    break
        else:
            self.extrair_relatorio(self)

    def reconhecer_itens_disputa(self):#CHAMA A FUNÇÃO DE ENVIO DE LANCES BASEADO NA ETAPA E TEMPO RESTANTE DE CADA ITEM
        print('reconhe_itens_disputa')
        #self.obter_botao_retirar_encerrados_disputa(self) botão de retirar encerrados acaba impedindo a continuação
        itens_em_disputa = self.obter_itens_tabela_disputa(self)
        for item in itens_em_disputa:
            item_disputa = self.obter_informacoes_item_disputa(self, item, self.obter_etapa_item_disputa(self,item))
            print(item_disputa['tempo_restante'])
            if(item_disputa['tempo_restante'] != tempo_prorrogado):
                print('tempo não prorrogado')
                if(item_disputa['tempo_restante']>5 and item_disputa['tempo_restante']<350):
                    print('decidir disputa do item: ',item_disputa['codigo_item'])
                    self.decidir_lance_tempo_normal_aberto(self,item_disputa)
    
    def decidir_lance_tempo_normal_aberto(self, item):
        for cotado in self.itens:
            if(item['codigo_item'] == cotado[0]):
                if(item['nosso_valor'] == item['atual_valor']):
                    #break
                #elif(item['atual_valor'] > cotado[3]):
                    print()
                    #decidir com base no intervalo mínimo entre lances
                else:
                    self.varrer_melhores_lances(self, item, cotado)
            break

    def varrer_melhores_lances(self, item, cotado):
        item['menu_lances'].click()
        try:
            item['webelement'].find_element_by_xpath('./div[3]/app-listagem-propostas-lances-item/p-tabview/div/ul/li[2]/a').click()
        except:
            item['menu_lances'].click()
            time.sleep(sel_delay)
            item['webelement'].find_element_by_xpath('./div[3]/app-listagem-propostas-lances-item/p-tabview/div/ul/li[2]/a').click()
        while(True):
            time.sleep(sel_delay)
            linhas_tabela = item['webelement'].find_elements_by_xpath('./div[3]/app-listagem-propostas-lances-item/p-tabview/div/div/p-tabpanel[2]/div/app-melhores-valores/div/div/table/tbody/*')
            if(len(linhas_tabela) > 0):
                print('linhas', len(linhas_tabela))
                break
        valores=[]
        for n in range(0,len(linhas_tabela)):
            aux_valor = linhas_tabela[n].find_element_by_xpath('./td[2]').text
            valores.append(float(converter_texto_para_decimal(aux_valor)))
            print('analisando o lance de ',aux_valor)
            print(valores)
            print(cotado)
            print(item)
            if((valores[n] > float(cotado[3])) and (valores[n] < float(item['nosso_valor']))):
                if((valores[n-1] < float(cotado[3])) and (valores[n]-1 < float(item['nosso_valor']))):
                    lance = valores[n]-item['intervalo_lance']
                    while(lance>float(item['nosso_valor'])-item['intervalo_lance']):
                        lance -= 0.05
                    print('Item ',item['item'],' -> R$ ',lance)

    # def decidir_lance(self,item):#DIFERENCIAR A ESTRATEGIA DE LANCE COM A ETAPA DO ITEM
    #     for cotado in self.itens:
    #         if(cotado[0] == item['codigo_item']):
    #             if(item['etapa_disputa'] == item_etapa_aberta):
    #                 atual = converter_texto_para_decimal(item['atual_valor'])
    #                 nosso = converter_texto_para_decimal(item['nosso_valor'])
    #                 intervalo = item['intervalo_lances']
    #                 item['menu_lances'].click()
    #                 time.sleep(sel_delay)
    #                 try:#ABRIR O TABELA COM OS MELHORES LANCES
    #                     item['webelement'].find_element_by_xpath('./div[3]/app-listagem-propostas-lances-item/p-tabview/div/ul/li[2]/a').click()
    #                     time.sleep(sel_delay)
    #                     item['webelement'].find_element_by_xpath('./div[3]/app-listagem-propostas-lances-item/p-tabview/div/ul/li[2]/a').click()
    #                 except:
    #                     item['webelement'].find_element_by_xpath('./div[2]/div[2]/div/app-botao-icone/span/button/i').click()
    #                     time.sleep(sel_delay)
    #                     item['webelement'].find_element_by_xpath('./div[3]/app-listagem-propostas-lances-item/p-tabview/div/ul/li[2]/a').click()
    #                 time.sleep(sel_delay)
    #                 linhas_tabela = item['webelement'].find_elements_by_xpath('./div[3]/app-listagem-propostas-lances-item/p-tabview/div/div/p-tabpanel[2]/div/app-melhores-valores/div/div/table/tbody/*')
    #                 valores = []
    #                 while(True):
    #                     time.sleep(sel_delay)
    #                     linhas_tabela = item['webelement'].find_elements_by_xpath('./div[3]/app-listagem-propostas-lances-item/p-tabview/div/div/p-tabpanel[2]/div/app-melhores-valores/div/div/table/tbody/*')
    #                     if(len(linhas_tabela) > 0):
    #                         break
    #                 for n in range(0,len(linhas_tabela)):
    #                     aux_valor = linhas_tabela[n].find_element_by_xpath('./td[2]').text
    #                     valores.append(float(converter_texto_para_decimal(aux_valor)))
    #                     print(aux_valor)
    #                     if((valores[n] > cotado[3]) and (valores[n] < float(nosso))):
    #                         if((valores[n-1] < cotado[3]) and (valores[n]-1 < float(nosso))):
    #                             lance = valores[n]-intervalo
    #                             while(lance>float(nosso)-intervalo):
    #                                 lance -= 0.05
    #                             print('Item ',item['item'],' -> R$ ',lance)
    #                             #enviar_lance(self,item,valor)
    #     item['menu_lances'].click()

    # def decidir_lance_fechado(self):
    #     atual = converter_texto_para_decimal(item['atual_valor'])
    #     nosso = converter_texto_para_decimal(item['nosso_valor'])
    #     intervalo = converter_intervalo_minimo(item['intervalo_lances'])
    #     if(cotado[3] < float(atual)):
    #         print('Nosso preço está mais baixo que o atual para o item ', item['item'])
    #         if(float(nosso) > (float(atual)*1.1)):
    #             print('Dar lance de R$ ', str(float(nosso)*1.1),' para o item ',item['item'])
    #             item['menu_lances'].click()
    #             return
    #         item['menu_lances'].click()#ABRIR E FECHAR MENU MELHORES LANCES
    #         return

    def enviar_lance(self, item, valor):
        valor = str(valor).replace('.',',')
        sel.enterFieldElement(item['input'],valor)
        item['botao_confirma'].click()
        sel.obter_elemento_xpath(self,'/html/body/modal-container/div/div/app-dialog-confirmacao/div/div/div[3]/div/div[2]/button').click()
        print('lance enviado')
        return

    def extrair_relatorio(self):#TESTAR TESTAR TESTAR
        print('Extraiir relatório')
        self.navegacao_itens[2].click()
        itens_encerrados = sel.obter_elementos_xpath(self,'/html/body/app-root/div/div/div/app-cabecalho-disputa-fornecedor/div[5]/div[2]/app-disputa-fornecedor/div/p-tabview/div/div/p-tabpanel[3]/div/app-disputa-fornecedor-itens/div/p-dataview/div/div[2]/div/div')
        print('Itens para relatório: ',len(itens_encerrados))
        resultado_por_item = []
        for item in itens_encerrados:
            aux_item = {}
            aux_item = {
                'item':item.find_element_by_xpath('./div[1]/div[1]/div/div/span[1]').text,
                'melhor_valor' : str(item.find_element_by_xpath('./div[2]/div[1]/div[2]/div/div[1]/div[2]/div[1]').text),
                'nosso_valor' : str(item.find_element_by_xpath('./div[2]/div[1]/div[2]/div/div[1]/div[2]/div[2]').text),
                }
            sel.clicar_subelemento(self.sel_driver,item,'./div[2]/div[2]/div/app-botao-icone/span/button/i')
            sel.clicar_subelemento(self.sel_driver,item,'./div[3]/app-listagem-propostas-lances-item/p-tabview/div/ul/li[2]/a')
            time.sleep(sel_delay)
            linhas_tabela = item.find_elements_by_xpath('./div[3]/app-listagem-propostas-lances-item/p-tabview/div/div/p-tabpanel[2]/div/app-melhores-valores/div/div/table/tbody/tr')
            while(True):
                time.sleep(sel_delay)
                linhas_tabela = item.find_elements_by_xpath('./div[3]/app-listagem-propostas-lances-item/p-tabview/div/div/p-tabpanel[2]/div/app-melhores-valores/div/div/table/tbody/tr')
                if(len(linhas_tabela)>0):
                    break
            colocacao = 1
            for linha in linhas_tabela:
                if (linha.find_element_by_xpath('./td[2]').text == aux_item['nosso_valor']):
                    break
                else:
                    colocacao += 1
            sel.clicar_subelemento(self.sel_driver, item, './div[2]/div[2]/div/app-botao-icone/span/button/i')
            aux_item['colocacao'] = colocacao
            resultado_por_item.append(aux_item)
        print(resultado_por_item)

bot = ComprasNet()
bot.iniciar()

def fechar_webdriver(self):
    choose = input('>')
    self.sel_driver.close()

fechar_webdriver(bot)