from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium import webdriver
import sel_operacoes_comum as sel
import openpyxl
import dados
import time
import sys
import os
from openpyxl import Workbook

excel_read=''

class andamento():
    
    def iniciar(self):
        self.verificar_planilha_controle()
        sel.configurar_webdriver(self,False)
        sel.coletar_credenciais_acessar_sistema(self)
        sel.acessar_menu_comprasnet(self)
        self.verificar_anexos()
        self.varrer_pregoes()
        self.finalizar()

    def verificar_planilha_controle(self):
        self.identidade = dados.identidade
        self.pregao_path = dados.pregao_path
        pregao_path = self.pregao_path
        try:
            self.excel_read = openpyxl.load_workbook(pregao_path)
        except:
            print("A planilha de controle de mensagens foi corrompida.")
            if os.path.exists(pregao_path):
                os.remove(pregao_path)
            wb = openpyxl.Workbook().save(pregao_path)
            self.excel_read = openpyxl.load_workbook(pregao_path)
            print("Planilha de controle recriada.")
    
    def verificar_anexos(self):
        sel.clicar_nova_janela_xpath(self,'/html/body/div[1]/ul/li[11]/a')
        time.sleep(0.2)
        windowToClose = self.sel_driver.window_handles[1]
        self.sel_driver.switch_to.window(windowToClose)
        table = sel.obter_elementos_xpath(self,'/html/body/div[1]/form/table/tbody/tr[5]/td[2]/table/tbody/*')
        del table[0]
        if(table[0].text == "No momento não existem pregões para enviar anexos."):
            print('Sem envio de anexos pendente')
        else:
            for rows in table:
                print('\nATENÇÃO É PRECISO ENVIAR ANEXOS PARA: ')
                info = rows.find_elements_by_xpath('./*')
                print("Pregão: "+info[1].text +"\nUASG: "+ info[2].text + "\nÓrgão: "+ info[3].text)
                print('\n')
        self.sel_driver.close()
        self.sel_driver.switch_to.window(self.sel_mainWindow)
    
    def varrer_pregoes(self):
        sel.trocar_frame(self,'/html/frameset/frameset/frame')
        sel.clicar_xpath(self,'/html/body/div[1]/ul/li[4]/a')

        nail = atualizar_lista_pregoes(self)

        print('No momento '+ str(len(nail)) +' pregões estão ativos.\n')

        #COMPARAR INFORMAÇÕES COM A PLANILHA
        for index in range(0,len(nail)):
            excel_read = self.excel_read
            atualizar_lista_pregoes(self)
            criar = True
            sheet_name=str(nail[index][0]+" | "+nail[index][1])
            for sheet in range(0,len(excel_read.sheetnames)):
                if(excel_read.worksheets[sheet].title == sheet_name):
                    criar = False
                    if(nail[index][3].text == 'Acompanhar'):
                        #try:
                        self.ler_pregao(linka = nail[index][3],sheet=sheet_name, exist=True)
                        #except:
                        #    print('Ops, aconteceu o erro ', sys.exc_info()[0])
                        #    print('Infelizmente não foi possível realizar a consulta ao pregão: nº ' + nail[index][0]+" | UASG: " + nail[index][1]+"\n")
            if(criar):
                excel_read.create_sheet(sheet_name)
                excel_read.save(self.pregao_path)
                if(nail[index][3].text == 'Acompanhar'):
                    for sheet in range(0,len(excel_read.sheetnames)):
                        if(excel_read.worksheets[sheet].title == sheet_name):
                            self.ler_pregao(nail[index][3], sheet_name, False)

    def ler_pregao(self, linka, sheet, exist):
        sel_mainWindow = self.sel_driver.window_handles[0]
        sel.clicar_nova_janela_elemento(self, linka)
        time.sleep(1)
        windowToClose = self.sel_driver.window_handles[1]
        self.sel_driver.switch_to.window(windowToClose)
        sel.clicar_xpath(self,'/html/body/div[1]/table/tbody/tr[6]/td[2]/input')
        self.sel_driver.close()
        sel_windowToClose = self.sel_driver.window_handles[1]
        self.sel_driver.switch_to.window(sel_windowToClose)
        table = sel.obter_elementos_xpath(self,'/html/body/table[2]/tbody/*')
        sh = self.excel_read[sheet]
        if(exist):
            msg = table[0].find_elements_by_xpath('./*')
            if(sh.cell(row=1, column=1).value == msg[0].text):
                print('_________________________ Processo: '+sheet+"\nSem novas atividades.\n")
            else:
                for index in range(0,len(table)):
                    if(index<50):
                        msg = table[index].find_elements_by_xpath('./*')
                        cl = sh.cell(row=index+1, column=1)
                        cl.value = msg[0].text
                        cl = sh.cell(row=index+1, column=2)
                        cl.value = msg[1].text
                        if self.identidade in msg[1].text:
                            print('_________________________ Processo: '+sheet)
                            print('Mensagem referenciando a sua indentidade foi enviada.\n'+msg[0].text+' expressando:\n'+msg[1].text)
                    if(index==0):
                        print('_________________________ Processo: '+sheet+" "+ msg[0].text + "\n" + msg[1].text+"\n")
        else:
            for index in range(0,len(table)):
                if(index<50):
                    msg = table[index].find_elements_by_xpath('./*')
                    cl = sh.cell(row=index+1, column=1)
                    cl.value = msg[0].text
                    cl = sh.cell(row=index+1, column=2)
                    cl.value = msg[1].text
                    if self.identidade in msg[1].text:
                        print('_________________________ Processo: '+sheet)
                        print('Mensagem referenciando a sua indentidade foi enviada.\n'+msg[0].text+' expressando:\n'+msg[1].text)
                if(index==0):
                        print('_________________________ Processo: '+sheet+" "+ msg[0].text + "\n" + msg[1].text+"\n")
        self.excel_read.save(self.pregao_path)
        self.sel_driver.close()
        self.sel_driver.switch_to.window(sel_mainWindow)
    
    def finalizar(self):
        print('Varredura finalizada.')
        self.sel_driver.quit()
        self.excel_read.save(self.pregao_path)
        self.excel_read.close()
        inp = input('>')
        sys.exit()
        return

def atualizar_lista_pregoes(self):
    sel.trocar_frame(self,'/html/frameset/frameset/frame')
    table = sel.obter_elementos_xpath(self,'/html/body/div[1]/table/tbody/tr[5]/td[2]/table/tbody/*')
    del table[0]
    def_nail = []
    for index in table:
        linha = index.find_elements_by_xpath('./*')
        aux_nail=[]
        aux_nail.append(linha[1].text)
        aux_nail.append(linha[2].text)
        aux_nail.append(linha[3].text)
        aux_nail.append(linha[0])
        def_nail.append(aux_nail)
    return def_nail

bot = andamento()
bot.iniciar()
exit()